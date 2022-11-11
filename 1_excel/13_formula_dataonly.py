from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)
        
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active
# evaluate 되지 않은 데이터는 None
for row in ws.values:
    for cell in row:
        print(cell)