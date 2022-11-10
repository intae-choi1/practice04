from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from random import randint
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0,100)])

col_B = ws["B"] # 영어(B column)만 가지고 오기
for cell in col_B:
    print(cell.value)
    
col_range = ws["B:C"] # 영어:수학
print(col_range)
for col in col_range:
    for cell in col:
        print(cell.value)

row_title = ws[1] # 1번쨰 row만 가지고 오기
for cell in row_title:
    print(cell.value)
    
row_range = ws[2:6] # 2,3,4,5,6 row (6포함)
for row in row_range:
    for cell in row:
        # coordinate로 좌표 get
        print(f"{cell.coordinate} {cell.value}", end=" ")
        xy = coordinate_from_string(cell.coordinate) # 좌표를 튜플로 나눠줌
        print(xy)
    print()
    

rows = tuple(ws.rows)   # row를 묶어서
cols = tuple(ws.columns)    # col을 묶어서
# print(cols)

# 1~5 row & 2~3 col 범위에서 row를 반복
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[1].value)

wb.save("sample.xlsx")