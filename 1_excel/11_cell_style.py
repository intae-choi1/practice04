from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# A 열의 너비를 5로
ws.column_dimensions["A"].width = 5

ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=15, underline="double")

# 테두리 적용
# None, 'dashDot','dashDotDot','dashed','dotted', 'double','hair','medium',
# 'mediumDashDot','mediumDashDotDot', 'mediumDashed','slantDashDot','thick','thin'
l_side = Side(style="thin")
r_side = Side(style="dashDotDot")
t_side = Side(style="dashDotDot")
b_side = Side(style="hair")
thin_border = Border(left=l_side, right=r_side, top=t_side, bottom=b_side)
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘으면 초록색 배경
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center") # left, right, top, bottom
        if cell.column == 1: continue # A 열은 제외
        if isinstance(cell.value, int) and cell.value > 87:
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid") # 배경을 초록색으로
            cell.font = Font(color="FF0000")
# 틀 고정
ws.freeze_panes = "B2"  # B2를 기준으로 고정 (위와 왼쪽)
wb.save("sample_style.xlsx")