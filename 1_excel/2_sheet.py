from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("yourSheet")
ws2 = wb.create_sheet("newSheet", 1) # 순서를 인덱스로 넣을 수 있음

print(wb.sheetnames)
new_ws = wb["newSheet"] # dict로 sheet에 접근가능

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied sheet"

wb.save("sample.xlsx")