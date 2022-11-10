from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active
ws.title = "mySheet"

# 셀에 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

print(ws["A1"].value)
print(ws["A10"].value)  # 빈곳은 null 반환

# col = A, B, C ,...
# row = 1, 2, 3, ...
a1 = ws.cell(column=1, row=1).value # ws["A1"]과 같다
print(a1)

c1 = ws.cell(column=3, row=1, value=10) # ws["C1"] = 10 과 같다
print(c1.value)


for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0,100))  # 0~100 숫자 랜덤 채우기

wb.save("sample.xlsx")