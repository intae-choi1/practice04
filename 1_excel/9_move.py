from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


ws.move_range("B1:C11", rows=0, cols=1) # B1:C11을 (rows=0, cols=1) 만큼 옮겨라

ws["B1"].value = "국어"

wb.save("sample_move.xlsx")