from openpyxl import Workbook
data = [
    ["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"],
    [1,10,8,5,14,26,12],
    [2,7,3,7,15,24,18],
    [3,9,5,8,8,12,4],
    [4,7,8,7,17,21,18],
    [5,7,8,7,16,25,15],
    [6,3,5,8,8,17,0],
    [7,4,9,10,16,27,18],
    [8,6,6,6,15,19,17],
    [9,10,10,9,19,30,19],
    [10,9,8,8,20,25,20]]

wb = Workbook()
ws = wb.active

for item in data:
    ws.append(item)

for i, cell in enumerate(ws['D']):
    if i == 0: continue
    cell.value = 9

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=9)):
    print(row)
    if i == 0: row[6].value = "총점"; row[7].value = "성적";continue
    row[6].value = f"=sum(B{i+1}:G{i+1})"

    total = sum([item.value for item in row[:6]])
    if total >= 90:
        row[7].value = "A"
    elif total >= 80:
        row[7].value = "B"
    elif total >= 70:
        row[7].value = "C"
    else: 
        row[7].value = "D"

    if row[0].value < 5:
        row[7].value = "F"



wb.save("quiz.xlsx")